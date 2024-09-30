from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from matplotlib.widgets import CheckButtons
import matplotlib.pyplot as plt
from matplotlib import patches
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import mplcursors
import subprocess
import pyautogui
import openpyxl
import logging
import time
import os
import re


def abrir_y_ejecutar_scanner(executable_path):
    # Inicia el programa
    os.startfile(executable_path)

    # Espera a que la aplicación se abra completamente
    time.sleep(1)

    # Automáticamente hace clic en el botón de "Escanear"
    scan_button_position = (85, 104)
    pyautogui.click(scan_button_position)

    # Muestra mensaje a que dia y hora, comenzo el escaneo
    dia = datetime.now().strftime("%d-%m-%Y")
    hora = datetime.now().strftime("%H:%M")
    print(f"El escaneo comenzó a las {dia} {hora}.")


def esperar_termino_scanner(file_path, imagen_boton, max_intentos=15, max_carpetas=10, tiempo_espera=30):
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
    logger = logging.getLogger(__name__)

    logger.info("Esperando a que termine el escaneo...")

    csv_file_path = file_path
    button_image_path = imagen_boton

    for intento in range(max_intentos):
        try:
            button7_location = pyautogui.locateOnScreen(
                button_image_path, confidence=0.9)
            if button7_location:
                logger.info(f"El escaneo terminó a las {datetime.now().strftime(
                    '%Y-%m-%d %H:%M:%S')}, procediendo a guardar el informe")

                if guardar_archivo(csv_file_path, max_carpetas):
                    return pyautogui.center(button7_location)
                else:
                    logger.error("Fallo al guardar el archivo.")
                    return None

            logger.info(f"No se encontró el botón. Intento {
                        intento + 1} de {max_intentos}.")
        except pyautogui.ImageNotFoundException:
            logger.warning(f"No se pudo encontrar la imagen. Intento {
                           intento + 1} de {max_intentos}.")

        if intento < max_intentos - 1:
            logger.info(
                f"Esperando {tiempo_espera} segundos antes del siguiente intento...")
            time.sleep(tiempo_espera)

    logger.error(
        "Se alcanzó el número máximo de intentos. No se pudo encontrar el botón.")
    return None


def guardar_archivo(csv_file_path, max_carpetas=10):
    logger = logging.getLogger(__name__)

    pyautogui.click(293, 619)
    # Configuración de botones
    buttons = [
        ((293, 619), "Tipo"),
        ((271, 703), "CSV"),
        ((341, 357), "archivo ip"),
        ((896, 710), "Guardar"),
        ((1024, 521), "Si")
    ]

    def click_button(coords, description):
        pyautogui.click(coords)
        logger.info(f"Click en {description}")

    def proceso_guardado():
        logger.info("Iniciando proceso de guardado")

        time.sleep(3)
        pyautogui.hotkey('ctrl', 's')

        for coords, description in buttons:
            click_button(coords, description)

        logger.info("Se inicia el guardado del archivo ip")

    # Primer intento de guardado
    proceso_guardado()

    # Verificación del archivo guardado
    for intento in range(max_carpetas):
        if os.path.exists(csv_file_path):
            modified_time = os.path.getmtime(csv_file_path)
            modified_datetime = datetime.fromtimestamp(modified_time)

            if (datetime.now() - modified_datetime).total_seconds() < 300:
                logger.info(f"Archivo '{csv_file_path}' guardado correctamente. "
                            f"Última modificación: {modified_datetime.strftime('%Y-%m-%d_%H-%M-%S')}")
                return True

            logger.warning(f"El archivo '{csv_file_path}' no se modificó en los últimos 5 minutos. "
                           f"Intentando de nuevo.")
        else:
            logger.warning(f"No se pudo guardar el archivo '{csv_file_path}'.")

        if intento < max_carpetas - 1:
            logger.info(f"Intentando guardar el archivo nuevamente. "
                        f"Intento {intento + 2} de {max_carpetas}.")
            # Intento adicional de hacer clic en los botones
            proceso_guardado()
            time.sleep(5)

    logger.error(
        "Se alcanzó el número máximo de intentos. No se pudo guardar el archivo.")
    return False


def cargar_datos(file_path):
    """Cargar datos desde un archivo CSV, manejando diferentes codificaciones."""
    encodings = ['utf-16', 'ISO-8859-1']
    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding,
                             delimiter='\t', on_bad_lines='skip')
            print(f"Datos cargados exitosamente del CSV.")
            return df
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"Error al cargar el archivo '{file_path}': {e}")
        return None


def calcular_conteos(df):
    """Generar gráficos de barras, circular y de barras apiladas para el estado de los dispositivos."""
    if df is None or 'Estado' not in df.columns:
        print("Columna 'Estado' no encontrada en el DataFrame.")
        return None
    # Calcular las variables
    conteo_estado = df['Estado'].value_counts()
    conteo_estado_porcentaje = df['Estado'].value_counts(normalize=True)

    df['Segmento'] = df['IP'].apply(lambda x: '.'.join(
        x.split('.')[:3]) + '.0/24' if pd.notna(x) else 'Desconocido')
    conteo_segmentos_estados = df.groupby(
        ['Segmento', 'Estado']).size().unstack(fill_value=0)

    # Mostrar las variables en la consola
    '''
    print("Conteo de Estados:")
    print(conteo_estado)
    print("\nPorcentaje de Estados:")
    print(conteo_estado_porcentaje)
    print("\nConteo de Estados por Segmento:")
    print(conteo_segmentos_estados)
    '''
    return conteo_estado, conteo_estado_porcentaje, conteo_segmentos_estados


def formatear_hoja(ws, num_rows):
    """Formatear una hoja de cálculo de Excel."""
  # Alinear los porcentajes a la derecha
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4, max_row=num_rows + 1):
        for cell in row:
            cell.alignment = Alignment(horizontal='right')

    # Alinear Totales y Porcentajes a la derecha
    for cell in ws['B'][num_rows - 2:]:
        cell.alignment = Alignment(horizontal='right')
    for cell in ws['B'][num_rows - 1:]:
        cell.alignment = Alignment(horizontal='right')

    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2


def exportar_a_excel(conteo_segmentos_estados, excel_path):
    """Exportar datos a un archivo Excel, colocando la fecha actual primero."""
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # Crear o cargar el archivo Excel
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    # Crear o cargar la hoja de "Conteos"
    ws_conteos = wb.create_sheet(title=f"Conteos_{timestamp}", index=0)

    # Crear el DataFrame combinado
    df_conteos = pd.DataFrame({
        'Segmento': conteo_segmentos_estados.index,
        'Activado': conteo_segmentos_estados['Activado'],
        'Inactivo': conteo_segmentos_estados['Inactivo'],
        'Desconocido': conteo_segmentos_estados['Desconocido']
    })

    # Añadir la fila de totales
    df_conteos.loc['Totales'] = df_conteos[[
        'Activado', 'Inactivo', 'Desconocido']].sum()

    # Calcular los porcentajes
    total_activado = df_conteos['Activado'].sum()
    total_inactivo = df_conteos['Inactivo'].sum()
    total_desconocido = df_conteos['Desconocido'].sum()

    total_general = total_activado + total_inactivo + total_desconocido

    # Asegurarse de que los porcentajes sean calculados solo si hay un total mayor que cero
    if total_general > 0:
        porcentajes = [
            f"{total_activado / total_general * 100:.2f}%",
            f"{total_inactivo / total_general * 100:.2f}%",
            f"{total_desconocido / total_general * 100:.2f}%"
        ]
    else:
        porcentajes = ["0%", "0%", "0%"]

    # Añadir la fila de porcentajes, asegurando que los valores están en la columna correcta
    df_conteos.loc['Porcentajes'] = [
        None, porcentajes[0], porcentajes[1], porcentajes[2]]

    # Asegurar que las filas 'Totales' y 'Porcentajes' tengan su texto correspondiente
    df_conteos.at['Totales', 'Segmento'] = 'Totales'
    df_conteos.at['Porcentajes', 'Segmento'] = 'Porcentajes'

    # Añadir los datos a la hoja de "Conteos"
    for r in dataframe_to_rows(df_conteos, index=False, header=True):
        ws_conteos.append(r)

    # Formatear la hoja de "Conteos"
    formatear_hoja(ws_conteos, len(df_conteos))

    # Guardar el archivo Excel
    wb.save(excel_path)
    print(f"Datos exportados exitosamente a {excel_path}")


def mostrar_graficos(conteo_estado, conteo_segmentos_estados):
    """Generar gráficos de barras, circular y de barras apiladas para el estado de los dispositivos."""

    estado_order = ['Desconocido', 'Inactivo', 'Activado']
    color_map = {'Desconocido': 'red',
                 'Inactivo': 'orange',
                 'Activado': 'green'}

    # Asignar los colores a cada estado
    colors = [color_map[estado] for estado in conteo_estado.index]

    # Crear la figura y los subgráficos
    fig, axs = plt.subplots(2, 2, figsize=(16, 14))

    # Gráfico de barras
    conteo_estado.plot(kind='bar', color=colors, ax=axs[0, 0])
    axs[0, 0].set_title('Conteo de Dispositivos por Estado')
    axs[0, 0].set_xlabel('')
    axs[0, 0].set_ylabel('Número de Dispositivos')
    axs[0, 0].set_xticklabels(conteo_estado.index, rotation=0)

    # Añadir etiquetas a las barras
    for i, v in enumerate(conteo_estado):
        axs[0, 0].text(i, v, str(v), ha='center', va='bottom', fontsize='7')

    # Gráfico circular
    conteo_estado.plot(kind='pie', autopct='%1.1f%%',
                       colors=colors, ax=axs[0, 1])
    axs[0, 1].set_title('Distribución de Dispositivos por Estado')
    axs[0, 1].set_ylabel('')

    # Gráfico de barras apiladas
    ax_stacked = fig.add_subplot(2, 1, 2)
    plt.subplots_adjust(hspace=0.4)  # Aumentar el espacio entre gráficos
    colors_segmento = [color_map[estado]
                       for estado in conteo_segmentos_estados.columns]
    bars = conteo_segmentos_estados.plot(
        kind='bar', stacked=True, ax=ax_stacked, color=colors_segmento)
    ax_stacked.set_title(
        'Conteo de Dispositivos por Segmento de IP y Estado', pad=10)
    ax_stacked.set_xlabel('Segmento de IP')
    ax_stacked.set_ylabel('Número de Dispositivos')
    ax_stacked.set_xticklabels(
        conteo_segmentos_estados.index, rotation=45, ha='right')
    ax_stacked.set_xlim(-0.5, len(conteo_segmentos_estados.index) - 0.5)

    # Ajustar los límites del eje y para dar más espacio arriba
    ylim = ax_stacked.get_ylim()
    ax_stacked.set_ylim(ylim[0], ylim[1] * 1.1)

    annot = ax_stacked.annotate("", xy=(0, 0), xytext=(20, 20), textcoords="offset points",
                                bbox=dict(boxstyle="round", fc="lightyellow",
                                          ec="orange", alpha=0.8, pad=0.5),
                                arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=.2", color="orange"))
    annot.set_visible(False)

    handles, labels = ax_stacked.get_legend_handles_labels()
    ordered_handles = [handles[labels.index(
        estado)] for estado in estado_order if estado in labels]
    ax_stacked.legend(ordered_handles, estado_order,
                      title='Estado', loc='upper left', bbox_to_anchor=(1, 1))

    def update_annot(bar, segmento, estado, valor):
        x = bar.get_x() + bar.get_width() / 2.
        y = bar.get_y() + bar.get_height() / 2.

        # Ajustar la posición vertical de la anotación
        # Si la barra está en el 70% superior del gráfico
        if y > ax_stacked.get_ylim()[1] * 0.7:
            xytext = (20, -20)  # Colocar la anotación debajo de la barra
        else:
            xytext = (20, 20)  # Colocar la anotación encima de la barra

        annot.xyann = xytext
        annot.xy = (x, y)
        text = f"{estado}: {valor}\n{segmento}"
        annot.set_text(text)

        # Ajustar el color de fondo según el estado
        if estado == 'Activado':
            annot.get_bbox_patch().set_facecolor('lightgreen')
        elif estado == 'Inactivo':
            annot.get_bbox_patch().set_facecolor('lightsalmon')
        else:  # 'Desconocido'
            annot.get_bbox_patch().set_facecolor('lightgray')

        annot.get_bbox_patch().set_alpha(0.9)

    def hover(event):
        vis = annot.get_visible()
        if event.inaxes == ax_stacked:
            for i, estado_bars in enumerate(bars.containers):
                for j, bar in enumerate(estado_bars):
                    if bar.contains(event)[0]:
                        segmento = conteo_segmentos_estados.index[j]
                        estado = conteo_segmentos_estados.columns[i]
                        valor = int(bar.get_height())
                        update_annot(bar, segmento, estado, valor)
                        annot.set_visible(True)
                        fig.canvas.draw_idle()
                        return
        if vis:
            annot.set_visible(False)
            fig.canvas.draw_idle()

    fig.canvas.mpl_connect("motion_notify_event", hover)

    fig.subplots_adjust(hspace=0.3)
    axs[1, 0].axis('off')
    axs[1, 1].axis('off')

    print(f"Mostrando Graficos de Datos Recientes.")
    plt.show()


def mostrar_grafico_historico(file_path):

    def leer_datos_desde_excel(archivo):
        xls = pd.ExcelFile(archivo)
        sheets = xls.sheet_names

        data_frames = []
        for sheet in sheets:
            # Extraer la fecha y hora del nombre de la hoja usando regex
            match = re.search(
                r'_(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})$', sheet)
            if match:
                fecha = match.group(1)
                hora = match.group(2).replace('-', ':')
                df = pd.read_excel(archivo, sheet_name=sheet)
                df['Fecha y Hora'] = pd.to_datetime(
                    f"{fecha} {hora}", format="%Y-%m-%d %H:%M:%S")
                data_frames.append(df)

        # Concatenar todos los DataFrames en uno solo
        all_data = pd.concat(data_frames, ignore_index=True)

        return all_data

    # Leer los datos desde el archivo Excel

    df = leer_datos_desde_excel(file_path)

    # Convertir 'Conteo' a enteros
    df['Conteo'] = pd.to_numeric(
        df['Conteo'], errors='coerce').fillna(0).astype(int)

    # Verificar si hay valores NaN en columnas importantes y limpiarlos
    df = df.dropna(subset=['Estado', 'Conteo'])

    # Agrupar por 'Fecha y Hora' y 'Estado' y sumar los conteos
    df_grouped = df.groupby(['Fecha y Hora', 'Estado'],
                            as_index=False).agg({'Conteo': 'sum'})

    # Crear el gráfico
    fig, ax = plt.subplots(figsize=(12, 6))

    # Ajusta los datos para el gráfico
    lines = []
    for estado in df_grouped['Estado'].unique():
        df_estado = df_grouped[df_grouped['Estado'] == estado]
        line, = ax.plot(df_estado['Fecha y Hora'],
                        df_estado['Conteo'], label=estado, marker='o')
        lines.append(line)

    plt.xlabel('Fecha y Hora')
    plt.ylabel('Conteo')
    plt.title('Equipos en la red')
    plt.legend()
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()

    # Crear la anotación que aparecerá al pasar el mouse
    annot = ax.annotate("", xy=(0, 0), xytext=(20, 20),
                        textcoords="offset points",
                        bbox=dict(boxstyle="round", fc="w"),
                        arrowprops=dict(arrowstyle="->"))
    annot.set_visible(False)

    # Función para actualizar la anotación
    def update_annot(line, ind):
        x, y = line.get_data()
        x_val = x[ind["ind"][0]]
        y_val = y[ind["ind"][0]]

        # Encuentra la fila en el DataFrame correspondiente a los datos de la línea
        df_estado = df_grouped[df_grouped['Estado'] == line.get_label()]
        fecha_hora = df_estado[df_estado['Fecha y Hora']
                               == x_val]['Fecha y Hora'].values[0]

        # Convertir numpy.datetime64 a datetime.datetime si es necesario
        if isinstance(fecha_hora, pd.Timestamp):
            fecha_hora = fecha_hora.to_pydatetime()
        elif isinstance(fecha_hora, np.datetime64):
            fecha_hora = pd.to_datetime(fecha_hora).to_pydatetime()

        # Formatear la fecha y hora en un formato más legible
        fecha_hora_formateada = fecha_hora.strftime("%Y-%m-%d %H:%M:%S")

        annot.xy = (x_val, y_val)
        text = f"{line.get_label()}\n{fecha_hora_formateada}\n{y_val}"
        annot.set_text(text)
        annot.get_bbox_patch().set_alpha(0.8)

    # Función para manejar los eventos del mouse
    def hover(event):
        vis = annot.get_visible()
        if event.inaxes == ax:
            for line in lines:
                cont, ind = line.contains(event)
                if cont:
                    update_annot(line, ind)
                    annot.set_visible(True)
                    fig.canvas.draw_idle()
                    return
        if vis:
            annot.set_visible(False)
            fig.canvas.draw_idle()

    fig.canvas.mpl_connect("motion_notify_event", hover)

    # Configurar los botones de selección
    # Posición en la parte superior del gráfico
    rax = plt.axes([0.4, 0.80, 0.2, 0.1])
    labels = df_grouped['Estado'].unique()
    visibility = [line.get_visible() for line in lines]
    check = CheckButtons(rax, labels, visibility)

    def func(label):
        for line in lines:
            if line.get_label() == label:
                line.set_visible(not line.get_visible())
        plt.draw()

    check.on_clicked(func)

    # Mostrar el gráfico
    print("Mostrando Gráficos Históricos.")
    plt.show()


def terminar_proceso(nombre_proceso):

    # Usar taskkill para cerrar el proceso
    def cerrar_proceso(nombre_proceso):
        print("Intentando cerrar el escaner...")
        try:
            result = subprocess.run(['taskkill', '/F', '/IM', nombre_proceso],
                                    stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            if result.returncode == 0:
                print(f"Proceso {nombre_proceso} cerrado exitosamente.")
            else:
                print(f"Error al intentar cerrar el proceso {nombre_proceso}:")
                print(result.stderr)
        except Exception as e:
            print(f"Ocurrió un error: {e}")

    cerrar_proceso(nombre_proceso)


def main():

    executable_path = r"C:\Program Files (x86)\Advanced IP Scanner\advanced_ip_scanner.exe"

    file_path = r'C:\Users\jvargas\Documents\ip.csv'
    excel_path = r"G:\Mi unidad\device_status_report.xlsx"
    imagen_boton = r'C:\Users\jvargas\Phyton\proceso_ip\btn.png'
    nombre_proceso = "advanced_ip_scanner.exe"

    abrir_y_ejecutar_scanner(executable_path)

    resultado_escaner = esperar_termino_scanner(file_path, imagen_boton)
    if resultado_escaner:

        def procesar_datos(file_path, excel_path):
            # Cargar datos desde el archivo CSV
            df = cargar_datos(file_path)
            if df is not None:
                # Mostrar gráficos actuales
                conteo_estado, conteo_estado_porcentaje, conteo_segmentos_estados = calcular_conteos(
                    df)

                # Exportar a Excel
                exportar_a_excel(conteo_segmentos_estados, excel_path)

                # Mostrar gráficos actuales
                mostrar_graficos(conteo_estado, conteo_segmentos_estados)

                # Mostrar gráficos historicos
                mostrar_grafico_historico(excel_path)
            else:
                print("No se pudieron cargar los datos.")

        # Llamar a la función de procesamiento
        procesar_datos(file_path, excel_path)

        print(f"Proceso completado exitosamente")
        terminar_proceso(nombre_proceso)

        # Calcular la próxima hora
        proxima_hora = datetime.now() + timedelta(hours=2)
        print()  # Espacio en blanco
        print(f"Siguiente Escaneo a las {
              proxima_hora.strftime('%H:%M')} horas.")

    else:
        print("El proceso no pudo completarse.")


if __name__ == "__main__":
    main()
